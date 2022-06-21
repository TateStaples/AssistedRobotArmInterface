# coding: utf-8
from kivy.app import App
from kivy.clock import Clock
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.image import Image
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.graphics import *

import numpy as np
from numpy.linalg import inv
import cv2
import sys
import time
import openpyxl
import math

import urx
from urx.robotiq_two_finger_gripper import Robotiq_Two_Finger_Gripper
from threading import Thread
import pyrealsense2 as rs

import threading

debugging = False
camera = True
robot = True
mac = False
logging = True

if not camera:
    video = cv2.VideoCapture(0)


class KinectView(App):
    main_flag = True

    # rgb format
    background_color = (225 / 255, 198 / 255, 153 / 255)
    button_color = (.5, .5, .5)
    text_color = (0, 0, 0)
    text_font = "Roboto"

    # movement - None = still
    running = True
    move_direction = None  # forward/backward/...
    rotate_direction = None  # rz+/rz-

    # ---------- log functions ---------- #

    def start_log(self, name="log.xlsx"):
        if os.path.exists(name):
            self.logbook = openpyxl.load_workbook(name)
            self.logsheet = self.logbook.worksheets[0]
            self.current_row = self.logsheet.max_row + 1
        else:
            self.logbook = openpyxl.workbook.Workbook()
            self.logsheet = self.logbook.worksheets[0]
            header = ["time", "rob x", "rob y", "rob z", "action id"]
            for col, item in enumerate(header, start=1):
                self.logsheet.cell(1, col, item)
            self.current_row = 2
        self.start_time = time.time()
        self.log_file = name
        self.log("Start")

    def function_wrapper(self, func):
        def fun(x):
            self.log(func.__name__)
            func()

        return fun

    def log_clock(self, dt):
        self.log("Clock")

    def end_log(self):
        self.log("End")
        self.save_log()

    def log(self, id):
        if not logging: return
        stuff = (time.time() - self.start_time,  # time
                 *(rob.getl()[:3] if robot else ('-', '-', '-')),  # position
                 id,  # action id
                 )
        for col, item in enumerate(stuff, start=1):
            self.logsheet.cell(self.current_row, col, item)
        self.current_row += 1

    def save_log(self):
        self.logbook.save(self.log_file)

    # ---------- util functions ---------- #
    @staticmethod
    def logDebug(*args):  # can remove return to log actions
        if debugging:
            print(args)

    @staticmethod
    def dimensions():  # the way mac does layout is different than windows / linux
        return int(Config.get('graphics', 'width')) * (2 if mac else 1), \
               int(Config.get('graphics', 'height')) * (2 if mac else 1)

    @staticmethod
    def pos_hint_from_tuple(tup):  # converts tuple to kivy format
        return {'x': tup[0], 'y': tup[1]}

    def make_button(self, press, text="", pos=(0, 0), size=None, font_size=20, background_normal=None, border=None,
                    release=None):
        if size is None: size = self.button_size_1
        pos = self.pos_hint_from_tuple(pos)
        if background_normal is None:  # if labeled box
            b = Button(text=text, pos_hint=pos, font_name=self.text_font, font_size=font_size, size=size,
                       size_hint=(None, None))
            b.background_color = self.button_color + (1,)
        else:  # if image
            b = Button(pos_hint=pos, size=size, size_hint=(None, None), background_normal=background_normal,
                       border=border)
        b.bind(on_press=self.function_wrapper(press))
        if release is not None:
            b.bind(on_release=self.function_wrapper(release))
        self.root.add_widget(b)
        return b

    def make_label(self, text, pos):
        font = self.title_font
        pos = self.pos_hint_from_tuple(pos)
        label = Label(text=text, pos_hint=pos, font_name=self.text_font, font_size=font, size_hint=(None, None),
                      halign="left",
                      color=self.text_color + (1,))
        label.bind(texture_size=label.setter('size'))
        self.root.add_widget(label)
        return label

    # ---------- init ---------- #
    def build(self):
        self.start_log()
        self.root = layout = FloatLayout()

        self.build_sizes()

        # ----- session control ----- #
        self.build_session_control()

        # ----- camera ----- #
        self.build_cams()

        # ----- nudge controls ----- #
        self.build_move()
        self.build_gripper()

        with layout.canvas.before:
            Color(self.background_color[0], self.background_color[1], self.background_color[2], 1)
            self.rect = Rectangle(
                size=(int(Config.get('graphics', 'width')) * 2, int(Config.get('graphics', 'height')) * 2),
                pos=layout.pos)

        Thread(target=self.move_thread).start()
        Thread(target=self.rotate_thread).start()
        # Thread(target=self.update_img).start()

        return layout

    def build_sizes(self):
        w, h = int(Config.get('graphics', 'width')), int(Config.get('graphics', 'height'))
        if not mac:
            w, h = w // 2, h // 2

        self.title_font = int(w / 15)

        img_size = i_w, i_h = (640, 480)
        self.main_img_size = img_size
        scale_factor = 1 / 2
        self.secondary_img_size = (int(i_w * scale_factor), int(i_h * scale_factor))

        self.button_size_1 = 1 / 6 * w, 1 / 6 * h
        self.return_button_size = self.button_size_1[0] * 2.25, self.button_size_1[1]
        self.session_button_size = self.button_size_1
        self.camera_button_size = self.button_size_1[0] * 1.5, self.button_size_1[1]
        self.elevation_button_size = self.button_size_1[0], self.button_size_1[1]
        self.gripper_button_size = self.elevation_button_size

        self.rotation_button_size = 1 / 6 * w, 1 / 4 * h
        self.left_right_size = 1 / 8 * w, 1 / 10 * h
        self.forward_backward_size = 1 / 10 * w, 1 / 10 * h

    def build_session_control(self):
        s_x, s_y = 0.05, 0.15
        self.session_label = self.make_label("Session Control", (s_x, s_y))
        self.session_start = self.make_button(self.on_session_start, text="Start", size=self.session_button_size,
                                              pos=(s_x, s_y - .1))
        self.session_stop = self.make_button(self.on_stop, text="Stop", size=self.session_button_size,
                                             pos=(s_x + .1, s_y - .1))

    def build_cams(self):
        self.establish_cameras()
        c_x, c_y = 0.2, 0.3
        self.switchCamButton = self.make_button(self.switch_camera, text="Switch View", pos=(c_x - .005, c_y - .03),
                                                size=self.camera_button_size)
        pos_main = self.pos_hint_from_tuple((0.05, c_y + 0.075))
        w, h = self.dimensions()
        i_w, i_h = self.main_img_size
        hint_w, hint_h = i_w / w, i_h / h
        left_hint = pos_main['x'] + hint_w
        pos_secondary = self.pos_hint_from_tuple((left_hint - (self.secondary_img_size[0] / w), 0.05))
        self.main_img = Image(on_touch_down=self.on_press_image, pos_hint=pos_main, size=self.main_img_size,
                              source="resource_manual/main.png", size_hint=(None, None))
        self.secondary_img = Image(pos_hint=pos_secondary, size=self.secondary_img_size,
                                   source="resource_manual/secondary.png", size_hint=(None, None))
        # self.root.bind(size=self._update_rect, pos=self._update_rect)
        self.root.add_widget(self.main_img)
        self.root.add_widget(self.secondary_img)
        Clock.schedule_interval(self.update_img, 1 / 60)

    def build_move(self):
        w, h = self.dimensions()
        border = (0, 0, 0, 0)
        far_left = self.main_img.pos_hint['x'] + self.main_img.size[0] / w
        m_x, m_y = far_left + 0.05, 0.4
        elevation_y = m_y + .1
        move_offset = .02

        self.armLabel = self.make_label("Arm Movement", (m_x, m_y + .51))
        self.returnButton = self.make_button(self.on_press_return, size=self.return_button_size,
                                             text="Go to start position", pos=(m_x + .02, m_y - .392 + .04))
        self.upButton = self.make_button(self.on_press_upButton, text="Lift", release=self.on_release_upBotton,
                                         pos=(m_x + move_offset, elevation_y + .06), size=self.elevation_button_size)
        self.downButton = self.make_button(self.on_press_downButton, release=self.on_release_downButton,
                                           text="Lower", pos=(m_x + move_offset + .1, elevation_y + .06),
                                           size=self.elevation_button_size)

        w1, h1 = self.left_right_size
        w2, h2 = self.forward_backward_size
        w, h = self.dimensions()
        d_long_x, d_short_x = w1 / w, w2 / w
        d_long_y, d_short_y = h2 / h, h1 / h

        left_right_y = m_y + .3
        forward_backward_x = m_x + .08 - d_short_x / 4

        self.leftButton = self.make_button(self.on_press_leftButton,
                                           background_normal="resources/left.png", size=self.left_right_size,
                                           pos=(forward_backward_x + move_offset - d_long_x, left_right_y + .06),
                                           border=border, release=self.on_release_leftButton)
        self.rightButton = self.make_button(self.on_press_rightButton,
                                            background_normal="resources/right.png", size=self.left_right_size,
                                            pos=(forward_backward_x + move_offset + d_short_x, left_right_y + .06),
                                            border=border, release=self.on_release_rightButton)
        self.backwardButton = self.make_button(self.on_press_backwardButton,
                                               background_normal="resources/backward.png",
                                               size=self.forward_backward_size,
                                               pos=(forward_backward_x + move_offset, left_right_y - d_long_y + .06),
                                               border=border, release=self.on_release_backwardButton)
        self.forwardButton = self.make_button(self.on_press_forwardButton,
                                              background_normal="resources/forward.png",
                                              size=self.forward_backward_size,
                                              pos=(forward_backward_x + move_offset, left_right_y + d_short_y + .06),
                                              border=border, release=self.on_release_forwardButton)

    def build_gripper(self):
        g_x, g_y = self.armLabel.pos_hint['x'], 0
        open_close_y = 0.13
        rotate_xoffset = .02
        gripper_button_size = self.button_size_1

        self.gripperLabel = self.make_label("Gripper Rotation", (g_x, g_y + 0.4 + .06))

        self.gripperButtonOpen1 = self.make_button(self.on_press_gripperButtonOpen, text="Open \nGripper",
                                                   size=gripper_button_size,
                                                   pos=(g_x + rotate_xoffset, open_close_y + .06))

        self.gripperButtonClose1 = self.make_button(self.on_press_gripperButtonClose, text="Close \nGripper",
                                                    size=gripper_button_size,
                                                    pos=(g_x + rotate_xoffset + .1, open_close_y + .06))

        width, _ = self.dimensions()
        far_right = self.gripperButtonClose1.size[0] / width + self.gripperButtonClose1.pos_hint['x']
        center_x = (g_x + far_right) / 2
        button_width = self.rotation_button_size[0] / width
        offset = 0.005

        rotate_y = 0.25
        border = (0, 0, 0, 0)
        self.rotateCounterClock = self.make_button(self.on_press_rotate_counter_clock,
                                                   release=self.on_release_rotate_counter_clock,
                                                   background_normal="resources/counter_clock.png",
                                                   size=self.rotation_button_size,
                                                   pos=(
                                                   center_x - button_width - offset + rotate_xoffset, rotate_y + .06),
                                                   border=border)
        self.rotateClock = self.make_button(self.on_press_rotate_clock, release=self.on_release_rotate_clock,
                                            background_normal="resources/clock.png", size=self.rotation_button_size,
                                            pos=(center_x + offset + rotate_xoffset, rotate_y + .06), border=border)

    def _update_rect(self, instance, value):
        self.rect.pos = instance.pos
        self.rect.size = instance.size

    def on_press_image(self, instance, touch):
        IMAGE_X, IMAGE_Y = self.main_img.pos
        view1_x, view1_y = self.main_img.size
        mouseX, mouseY = self.main_img.to_widget(*touch.pos)
        self.logDebug("Mouse Click: ", mouseX, mouseY)
        if IMAGE_X < mouseX < IMAGE_X + view1_x and IMAGE_Y < mouseY < IMAGE_Y + view1_y:
            Thread(target=self.waypoint_thread, args=[mouseX, mouseY]).start()
        else:
            self.logDebug("Mouse click out of range")

    def waypoint_thread(self, mouseX, mouseY):
        IMAGE_X, IMAGE_Y = self.main_img.pos
        view1_x, view1_y = self.main_img.size
        c, r = mouseX - IMAGE_X, view1_y - (mouseY - IMAGE_Y)
        # print("$$$$$$$$$", c, r)
        depth_value = depth_image[r][c] * self.get_depth_scale()
        depth_point = rs.rs2_deproject_pixel_to_point(aligned_depth_intrinsics, [c, r], depth_value)
        camera_coordinates = np.array(depth_point)
        robot_coordinates = np.dot(R, camera_coordinates.T).T + t
        r_x, r_y, r_z = robot_coordinates[0], robot_coordinates[1], robot_coordinates[2]
        rob.movel((r_x, r_y, r_z, 3, -1, 0), a, v)
        time.sleep(0.2)

    def on_session_start(self):
        Clock.schedule_interval(self.log_clock, 1 / 10)
        Clock.schedule_interval(self.save_log, 2)  # save the log every 2 seconds in case of error

    def on_stop(self):
        self.logDebug("Stopping")
        self.running = False
        if robot: rob.close()
        self.end_log()
        sys.exit()

    # ---------- main ---------- #
    def on_press_pauseButton(self):
        Thread(target=self.pause_thread).start()

    def pause_thread(self):
        rob.stopl()

    def on_press_startButton(self):
        Thread(target=self.start_thread).start()

    def start_thread(self):
        rob.stopl()

    def on_press_stopButton(self):
        Thread(target=self.stop_thread).start()

    def stop_thread(self):
        rob.stopl()

    def establish_cameras(self, new=True):
        if camera:
            if new:
                self.pipeline1 = rs.pipeline()
                self.pipeline2 = rs.pipeline()
            profile1 = self.pipeline1.start(  # outer camera
                create_config(device_1, self.main_img_size if self.main_flag else self.secondary_img_size,
                              self.main_flag))
            profile2 = self.pipeline2.start(  # arm camera
                create_config(device_2, self.main_img_size if not self.main_flag else self.secondary_img_size,
                              not self.main_flag))
            self.depth_scale = (
                profile1 if self.main_flag else profile2).get_device().first_depth_sensor().get_depth_scale()

    def switch_camera(self):
        self.logDebug("Swapping cameras")
        self.main_flag = not self.main_flag
        if camera:
            self.pipeline1.stop()
            self.pipeline2.stop()
        self.establish_cameras(new=False)

    msg_time = time.time() + 10
    msg = "test"
    border_color = None  # clear

    @staticmethod
    def camera_mask(img):
        if not camera: return img
        alpha = .3

        table_height = -.2
        depth = rob.getl()[1] - table_height if robot else 1

        w, h, d = img.shape
        overlay = create_mask(depth, (w, h))

        overlay = cv2.resize(overlay, (h, w))
        mask = overlay[:, :, 3] / 255.0
        mask = cv2.merge((mask, mask, mask))
        overlay = cv2.cvtColor(overlay.astype('uint8'), cv2.COLOR_BGRA2BGR)

        img = img.astype('uint8')

        adjusted_overlay = overlay * alpha
        adjusted_img = img * (1 - mask * alpha)
        final_img = adjusted_img + adjusted_overlay
        # cv2.addWeighted(overlay, alpha, img, 1-alpha, 0, img)
        return final_img

    def add_notice(self, msg, t=3):
        self.msg = msg
        self.msg_time = time.time() + t

    def draw_notices(self, img):
        border_size = 10
        if self.border_color is not None:
            cv2.rectangle(img, (0, 0), self.main_img_size, color=self.border_color, thickness=border_size)
        if self.msg is not None:
            img = cv2.putText(img, str(self.msg), (10, 50), cv2.cv2.FONT_HERSHEY_PLAIN, 2.5, (0, 0, 200), 3)
            if time.time() > self.msg_time:
                self.msg = None
        return img

    def nudge_error(self):
        self.border_color = (0, 0, 255)
        self.add_notice("Reached workspace edge")

    def nudge_close(self):
        self.border_color = (0, 255, 255)
        self.add_notice("Close to workspace edge")

    # ---------- gripper ---------- #
    def on_press_gripperButtonOpen(self):
        Thread(target=self.gripper_thread_open).start()

    def on_press_gripperButtonClose(self):
        Thread(target=self.gripper_thread_close).start()

    def gripper_thread_open(self):
        robotiqgrip.open_gripper()

    def gripper_thread_close(self):
        robotiqgrip.close_gripper()

    # ---------- movement ---------- #
    def on_press_return(self):
        Thread(target=self.home_thread).start()

    def home_thread(self):
        rob_x, rob_y, rob_z, rob_rx, rob_ry, rob_rz = rob.getl()
        if rob_z < 0.21:
            rob.movels([(rob_x, rob_y, rob_z + 0.09, rob_rx, rob_ry, rob_rz), (0.5, 0.4, 0.3, 3, -1, 0)], a, v)
        else:
            rob.movel((0.5, 0.4, 0.3, 3, -1, 0), a, v)

        time.sleep(0.2)

    def on_press_upButton(self):
        self.upButton.disabled = True
        self.border_color = None
        Thread(target=self.move_thread, args=["up"]).start()
        time.sleep(0.1)

    def on_release_upBotton(self):
        if robot: rob.stopl()
        try:
            self.timer_close.cancel()
            self.timer_limit.cancel()
        except:
            pass
        self.border_color = None
        self.upButton.disabled = False
        Thread(target=self.move_thread, args=["up"])._stop

    def on_press_downButton(self):
        self.downButton.disabled = True
        self.border_color = None
        Thread(target=self.move_thread, args=["down"]).start()
        time.sleep(0.1)

    def on_release_downButton(self):
        if robot: rob.stopl()
        try:
            self.timer_close.cancel()
            self.timer_limit.cancel()
        except:
            pass
        self.border_color = None
        self.downButton.disabled = False
        Thread(target=self.move_thread, args=["down"])._stop

    def on_press_leftButton(self):
        self.leftButton.disabled = True
        self.border_color = None
        Thread(target=self.move_thread, args=["left"]).start()
        time.sleep(0.1)

    def on_release_leftButton(self):
        if robot: rob.stopl()
        try:
            self.timer_close.cancel()
            self.timer_limit.cancel()
        except:
            pass
        self.border_color = None
        self.leftButton.disabled = False
        Thread(target=self.move_thread, args=["left"])._stop

    def on_press_rightButton(self):
        self.rightButton.disabled = True
        self.border_color = None
        Thread(target=self.move_thread, args=["right"]).start()
        time.sleep(0.1)

    def on_release_rightButton(self):
        if robot: rob.stopl()
        try:
            self.timer_close.cancel()
            self.timer_limit.cancel()
        except:
            pass
        self.border_color = None
        self.rightButton.disabled = False
        Thread(target=self.move_thread, args=["right"])._stop

    def on_press_backwardButton(self):
        self.backwardButton.disabled = True
        self.border_color = None
        Thread(target=self.move_thread, args=["backward"]).start()
        time.sleep(0.1)

    def on_release_backwardButton(self):
        if robot: rob.stopl()
        try:
            self.timer_close.cancel()
            self.timer_limit.cancel()
        except:
            pass
        self.border_color = None
        self.backwardButton.disabled = False
        Thread(target=self.move_thread, args=["backward"])._stop

    def on_press_forwardButton(self):
        self.forwardButton.disabled = True
        self.border_color = None
        Thread(target=self.move_thread, args=["forward"]).start()
        time.sleep(0.1)

    def on_release_forwardButton(self):
        if robot: rob.stopl()
        try:
            self.timer_close.cancel()
            self.timer_limit.cancel()
        except:
            pass
        self.border_color = None
        self.forwardButton.disabled = False
        Thread(target=self.move_thread, args=["forward"])._stop

    # ---------- rotation ---------- #
    def on_press_rotate_clock(self):
        self.rotateClock.disabled = True
        self.border_color = None
        Thread(target=self.rotate_thread, args=["rz-"]).start()
        time.sleep(0.1)

    def on_release_rotate_clock(self):
        if robot: rob.stopl()
        try:
            self.timer_close.cancel()
            self.timer_limit.cancel()
        except:
            pass
        self.border_color = None
        self.rotateClock.disabled = False
        Thread(target=self.rotate_thread, args=["rz-"])._stop

    def on_press_rotate_counter_clock(self):
        self.rotateCounterClock.disabled = True
        self.border_color = None
        Thread(target=self.rotate_thread, args=["rz+"]).start()
        time.sleep(0.1)

    def on_release_rotate_counter_clock(self):
        if robot: rob.stopl()
        self.border_color = None
        try:
            self.timer_close.cancel()
            self.timer_limit.cancel()
        except:
            pass
        self.rotateCounterClock.disabled = False
        Thread(target=self.rotate_thread, args=["rz+"])._stop

    # ---------------------------------------------------------------------------------#

    def move_thread(self, direction):  # this probably needs updating
        if not robot:
            return
        # rob_x, rob_y, rob_z, rob_rx, rob_ry, rob_rz = rob.getl()
        # rob_pos = np.array([rob_x, rob_y, rob_z])
        # camera_pos = np.dot(inversed_R_back, rob_pos.T).T + inversed_t_back
        # cam_x, cam_y, cam_z = camera_pos

        Transform_EE2BASE_move = rob.get_pose()

        T_EE2BASE_move = Transform_EE2BASE_move.array
        R_EE2BASE_move = np.array([[T_EE2BASE_move[0, 0], T_EE2BASE_move[0, 1], T_EE2BASE_move[0, 2]],
                                   [T_EE2BASE_move[1, 0], T_EE2BASE_move[1, 1], T_EE2BASE_move[1, 2]],
                                   [T_EE2BASE_move[2, 0], T_EE2BASE_move[2, 1], T_EE2BASE_move[2, 2]]])
        # translation
        Tr_EE2BASE_move = np.array([T_EE2BASE_move[0, 3], T_EE2BASE_move[1, 3], T_EE2BASE_move[2, 3]])

        T_BASE2EE_move = inv(T_EE2BASE_move)
        # rotation
        R_BASE2EE_move = np.array([[T_BASE2EE_move[0, 0], T_BASE2EE_move[0, 1], T_BASE2EE_move[0, 2]],
                                   [T_BASE2EE_move[1, 0], T_BASE2EE_move[1, 1], T_BASE2EE_move[1, 2]],
                                   [T_BASE2EE_move[2, 0], T_BASE2EE_move[2, 1], T_BASE2EE_move[2, 2]]])
        # translation
        Tr_BASE2EE_move = np.array([T_BASE2EE_move[0, 3], T_BASE2EE_move[1, 3], T_BASE2EE_move[2, 3]])

        rob_x, rob_y, rob_z, rob_rx, rob_ry, rob_rz = rob.getl()
        rob_pos_in_base = np.array([rob_x, rob_y, rob_z])
        rob_pos_in_EE_urx = np.dot(R_BASE2EE_move, rob_pos_in_base.T).T + Tr_BASE2EE_move
        rob_pos_EE_k = np.dot(R_k2ur5_inv, rob_pos_in_EE_urx.T).T
        rob_pos_in_cam = np.dot(R_on_arm_inv, rob_pos_EE_k.T).T + t_on_arm_inv

        cam_x, cam_y, cam_z = rob_pos_in_cam

        MOVE_UNIT = 100
        dist_limit = .85

        if direction == "up":
            MOVE_UNIT = 0.4 - rob_z
            camera_pos = np.array([cam_x, cam_y, cam_z - MOVE_UNIT])

        if direction == "down":
            MOVE_UNIT = rob_z + .15
            camera_pos = np.array([cam_x, cam_y, cam_z + MOVE_UNIT])

        if direction == "left":
            camera_pos = np.array([cam_x - MOVE_UNIT, cam_y, cam_z])

            rob_target_EE_K = np.dot(R_on_arm, camera_pos.T).T + t_on_arm
            rob_target_EE_urx = np.dot(R_k2ur5, rob_target_EE_K.T).T
            rob_target_base = np.dot(R_EE2BASE_move, rob_target_EE_urx.T).T + Tr_EE2BASE_move
            r_x, r_y, r_z = rob_target_base[0], rob_target_base[1], rob_target_base[2]
            target_dir = np.array([r_x - rob_x, r_y - rob_y, r_z - rob_z])

            initial_dist = (target_dir[0] ** 2 + target_dir[1] ** 2 + target_dir[2] ** 2) ** .5
            theta = np.lib.scimath.arccos(
                np.dot(target_dir, rob_target_base) / (np.linalg.norm(target_dir) * np.linalg.norm(rob_target_base)))

            root_x = np.linalg.norm(rob_target_base)
            delta = ((2 * root_x * np.cos(theta)) ** 2 - 4 * (root_x ** 2 - dist_limit ** 2)) ** .5  # sqrt(b**2-4ac)
            root_b = 2 * root_x * np.cos(theta)

            root_1 = (root_b + delta) * .5
            root_2 = (root_b - delta) * .5
            subtracts = np.sort([root_1, root_2])
            subtract = subtracts[0]
            MOVE_UNIT = initial_dist - subtract
            if MOVE_UNIT < 0:
                MOVE_UNIT = 0

            camera_pos = np.array([cam_x - MOVE_UNIT, cam_y, cam_z])

        if direction == "right":
            camera_pos = np.array([cam_x + MOVE_UNIT, cam_y, cam_z])

            rob_target_EE_K = np.dot(R_on_arm, camera_pos.T).T + t_on_arm
            rob_target_EE_urx = np.dot(R_k2ur5, rob_target_EE_K.T).T
            rob_target_base = np.dot(R_EE2BASE_move, rob_target_EE_urx.T).T + Tr_EE2BASE_move
            r_x, r_y, r_z = rob_target_base[0], rob_target_base[1], rob_target_base[2]
            target_dir = np.array([r_x - rob_x, r_y - rob_y, r_z - rob_z])

            initial_dist = (target_dir[0] ** 2 + target_dir[1] ** 2 + target_dir[2] ** 2) ** .5
            theta = np.lib.scimath.arccos(
                np.dot(target_dir, rob_target_base) / (np.linalg.norm(target_dir) * np.linalg.norm(rob_target_base)))

            root_x = np.linalg.norm(rob_target_base)
            delta = ((2 * root_x * np.cos(theta)) ** 2 - 4 * (root_x ** 2 - dist_limit ** 2)) ** .5  # sqrt(b**2-4ac)
            root_b = 2 * root_x * np.cos(theta)

            root_1 = (root_b + delta) * .5
            root_2 = (root_b - delta) * .5
            subtracts = np.sort([root_1, root_2])
            subtract = subtracts[0]
            MOVE_UNIT = initial_dist - subtract
            if MOVE_UNIT < 0:
                MOVE_UNIT = 0

            camera_pos = np.array([cam_x + MOVE_UNIT, cam_y, cam_z])

        if direction == "forward":
            camera_pos = np.array([cam_x, cam_y - MOVE_UNIT, cam_z])

            rob_target_EE_K = np.dot(R_on_arm, camera_pos.T).T + t_on_arm
            rob_target_EE_urx = np.dot(R_k2ur5, rob_target_EE_K.T).T
            rob_target_base = np.dot(R_EE2BASE_move, rob_target_EE_urx.T).T + Tr_EE2BASE_move
            r_x, r_y, r_z = rob_target_base[0], rob_target_base[1], rob_target_base[2]
            target_dir = np.array([r_x - rob_x, r_y - rob_y, r_z - rob_z])

            initial_dist = (target_dir[0] ** 2 + target_dir[1] ** 2 + target_dir[2] ** 2) ** .5
            theta = np.lib.scimath.arccos(
                np.dot(target_dir, rob_target_base) / (np.linalg.norm(target_dir) * np.linalg.norm(rob_target_base)))

            root_x = np.linalg.norm(rob_target_base)
            delta = ((2 * root_x * np.cos(theta)) ** 2 - 4 * (root_x ** 2 - dist_limit ** 2)) ** .5  # sqrt(b**2-4ac)
            root_b = 2 * root_x * np.cos(theta)

            root_1 = (root_b + delta) * .5
            root_2 = (root_b - delta) * .5
            subtracts = np.sort([root_1, root_2])
            subtract = subtracts[0]
            MOVE_UNIT = initial_dist - subtract
            if MOVE_UNIT < 0:
                MOVE_UNIT = 0

            camera_pos = np.array([cam_x, cam_y - MOVE_UNIT, cam_z])

        if direction == "backward":
            MOVE_UNIT = (rob_x ** 2 + rob_y ** 2) ** .5 - .3
            if MOVE_UNIT < 0:
                MOVE_UNIT = 0
            camera_pos = np.array([cam_x, cam_y + MOVE_UNIT, cam_z])

        # rob_pos = np.dot(R_back, camera_pos.T).T + t_back
        rob_target_EE_K = np.dot(R_on_arm, camera_pos.T).T + t_on_arm
        rob_target_EE_urx = np.dot(R_k2ur5, rob_target_EE_K.T).T
        rob_target_base = np.dot(R_EE2BASE_move, rob_target_EE_urx.T).T + Tr_EE2BASE_move
        r_x, r_y, r_z = rob_target_base[0], rob_target_base[1], rob_target_base[2]
        rob_rx, rob_ry, rob_rz = rob.getl()[3:6]
        self.logDebug(rob_target_base)

        move_time = MOVE_UNIT / v

        if MOVE_UNIT > .05:
            self.timer_close = threading.Timer(move_time + 3, self.nudge_close)
            self.timer_limit = threading.Timer(move_time + 5.5, self.nudge_error)
            self.timer_close.start()
            self.timer_limit.start()
        else:
            self.nudge_error()
            # time.sleep(1)
        rob.movel((r_x, r_y, r_z, rob_rx, rob_ry, rob_rz), a, v)

        # self.nudge_error()

        self.border_color = None

    def rotate_thread(self, direction):
        rob_o = rob.get_orientation()
        rob_pos = rob.getl()
        rob_rx, rob_ry, rob_rz = rob_pos[3], rob_pos[4], rob_pos[5]

        if direction == "rz+":
            rob_o.rotate_zb(rob_rz + ROTATE_UNIT)
            rob.set_orientation(rob_o, acc=0.03, vel=0.05)

        elif direction == "rz-":
            rob_o.rotate_zb(rob_rz - ROTATE_UNIT)
            rob.set_orientation(rob_o, acc=0.03, vel=0.05)

    def update_img(self, *args):
        # kivy's recommendation for video streaming:
        # https://kivyapps.wordpress.com/video-streaming-using-kivy-and-python/
        if not camera:
            ret, outer = video.read()
            arm = np.ones(self.secondary_img_size + (3,)) * 150
        else:
            # Get frameset of color and depth
            frames1 = self.pipeline1.wait_for_frames()
            frames2 = self.pipeline2.wait_for_frames()
            # frames.get_depth_frame() is a 640x360 depth image

            # Align the depth frame to color frame
            aligned_frames1 = align.process(frames1)
            aligned_frames2 = align.process(frames2)

            # Get aligned frames
            color_frame1 = aligned_frames1.get_color_frame()
            color_frame2 = aligned_frames2.get_color_frame()

            color_image1 = np.asanyarray(color_frame1.get_data())
            color_image2 = np.asanyarray(color_frame2.get_data())

            aligned_depth_frame = aligned_frames1.get_depth_frame() if self.main_flag else aligned_frames2.get_depth_frame()
            self.aligned_depth_intrinsics = rs.video_stream_profile(aligned_depth_frame.profile).get_intrinsics()
            self.depth_image = np.asanyarray(aligned_depth_frame.get_data())

            outer = color_image1
            arm = color_image2
            # cv2.imshow("image 1", color_image1)

        arm = self.camera_mask(arm)

        frame_big = outer if self.main_flag else arm  # cv2.resize(outer if self.main_flag else arm, self.main_img_size)
        frame_small = outer if not self.main_flag else arm  # cv2.resize(outer if not self.main_flag else arm, self.secondary_img_size)

        frame_big = self.draw_notices(frame_big)
        cv2.imwrite("resource_manual/main.png", frame_big)
        cv2.imwrite("resource_manual/secondary.png", frame_small)
        self.main_img.reload()
        self.secondary_img.reload()


def create_mask(depth, shape):
    h, w = shape
    img = np.zeros((h, w, 4))
    red = (0, 0, 255, 255)  # bgr

    # camera shadow
    fov_h = 64  # 86 - https://www.intel.com/content/www/us/en/support/articles/000030385/emerging-technologies/intel-realsense-technology.html
    fov_v = 41  # 57

    cam_width = .098
    cam_height = .024


    # camera stuff
    shadow_theta_h = math.degrees(math.atan(cam_width / 2 / depth))
    shadow_theta_v = math.degrees(math.atan(cam_height / 2 / depth))

    shadow_w = w * (shadow_theta_h / fov_h)
    shadow_h = h * (shadow_theta_v / fov_v)
    cx, cy = w // 2, h // 2
    cv2.rectangle(img, (int(cx - shadow_w), int(cy - shadow_h)), (int(cx + shadow_w), int(cy + shadow_h)), color=red,
                  thickness=-1)

    # gripper stuff
    gripper_width = 0.11
    gripper_length = 0.022
    offset = 0.09
    # cam_to_grip = .135

    cam_dh = int(w * math.degrees(math.atan(gripper_width / 2 / depth)) / fov_h)
    cam_dv = int(h * math.degrees(math.atan(gripper_length / depth)) / fov_v)
    offset_v = int(h * math.degrees(math.atan(offset / depth)) / fov_v)
    width_h = 10

    cv2.rectangle(img, (cx-cam_dh, cy-offset_v), (cx - cam_dh - width_h, cy-offset_v-cam_dv), red, -1)
    cv2.rectangle(img, (cx+cam_dh, cy-offset_v), (cx + cam_dh + width_h, cy-offset_v-cam_dv), red, -1)
    # cv2.imwrite("test.png", img)
    return img


def create_config(device, size, depth):
    config = rs.config()
    config.enable_device(device)
    if depth:
        config.enable_stream(rs.stream.depth, size[0], size[1], rs.format.z16, 30)
    config.enable_stream(rs.stream.color, size[0], size[1], rs.format.bgr8, 30)
    return config


import os

os.environ['KIVY_GL_BACKEND'] = 'sdl2'
from kivy.config import Config

# 0 being off 1 being on as in true / false
# you can use 0 or 1 && True or False
Config.set('graphics', 'resizable', '0')

# fix the width of the window
# was 1500 * 900
Config.set('graphics', 'width', '500' if mac else '1000')
Config.set('graphics', 'height', '400' if mac else '800')

device_1 = '911222060024'
device_2 = '916512060628'

# acceleration and velocity
v = .5
a = 0.04

# start the robot arm
if not robot:
    rob = None
    robotiqgrip = None
else:
    # """
    rob = urx.Robot("192.168.1.105")
    rob.set_tcp((0, 0, 0.18, 0, 0, 0))
    time.sleep(0.2)
    robotiqgrip = Robotiq_Two_Finger_Gripper(rob)
    # """

# Create an align object
# rs.align allows us to perform alignment of depth frames to others frames
# The "align_to" is the stream type to which we plan to align depth frames.
if camera:
    align_to = rs.stream.color
    align = rs.align(align_to)

# R = np.array([[0.4729613, -0.02498102, 0.88072899],
#               [-0.88017677, 0.03193417, 0.47357054],
#               [-0.03995563, -0.99917774, -0.00688409]])
# t = np.array([-0.43179749, 0.24159906, 0.15602798])

# inversed_R = np.array([[0.4729613, -0.88017677, -0.03995563],
#                        [-0.02498102, 0.03193417, -0.99917774],
#                        [0.88072899, 0.47357054, -0.00688409]])
# inversed_t = np.array([0.42310757, 0.13739768, 0.26695648])


R = np.array([[0.4729613, -0.02498102, 0.88072899],
              [-0.88017677, 0.03193417, 0.47357054],
              [-0.03995563, -0.99917774, -0.00688409]])
t = np.array([-0.43179749, 0.29159906, 0.15602798])

inversed_R = np.array([[0.472961306628662, -0.880176777913776, -0.0399556246279741],
                       [-0.0249810255046716, 0.0319341736638282, -0.999177741061022],
                       [0.880728990275949, 0.473570537677987, -0.00688408697281732]])
inversed_t = np.array([0.467116421543204, 0.135800965465922, 0.243277953925307])

# transformation from camera space to end effector space, this end effector is from klampt
R_on_arm = np.array([[0.0330604524021490, -0.170386879727978, 0.984822480299637],
                     [-0.998170445145012, 0.0442810919530209, 0.0411697380907021],
                     [-0.0506237980200431, -0.984381783716060, -0.168611194652118]])
t_on_arm = np.array([-0.00032808, -0.01807384, 0.09210034])

R_on_arm_inv = np.array([[0.03306045240214919338, -0.99817044514501170206, -0.050623798020043130422],
                         [-0.170386879727978593, 0.04428109195302085153, -0.98438178371605963418],
                         [0.98482248029963678239, 0.041169738090701849354, -0.16861119465211731576]])

t_on_arm_inv = np.array([-0.013367457435318322066, 0.09140632581353858694, 0.01659634417369614205])

# transformation matrix from klampt definition to ur5 default
R_k2ur5 = np.array([[0, -1, 0], [0, 0, -1], [1, 0, 0]])
R_k2ur5_inv = inv(R_k2ur5)

MOVE_UNIT = .5
ROTATE_UNIT = 3

app = KinectView()
app.run()
app.end_log()
cv2.destroyAllWindows()

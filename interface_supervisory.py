# coding: utf-8
from kivy.app import App
from kivy.uix.slider import Slider
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.image import Image
from kivy.clock import Clock
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.graphics import *

import numpy as np
from numpy.linalg import inv
import cv2
import sys
import time
import math
import openpyxl

import urx
from urx.robotiq_two_finger_gripper import Robotiq_Two_Finger_Gripper
from threading import Thread, Lock
import pyrealsense2 as rs

# doesnt actually connect
mac = True
debugging = False
camera = False
robot = False
logging = False

if not camera:
    video = cv2.VideoCapture(0)


def good_round(num, to):
    return round(num / to) * to


class KinectView(App):
    # if the main camera is the back camera
    back_main_flag = True
    waypoint = None
    waypoint_pixel = None
    running = True
    delete_flag = False

    # rgb format
    background_color = (225 / 255, 198 / 255, 153 / 255)  # (255 / 255, 255 / 255, 255 / 255)
    button_color = (0.5, 0.5, 0.5)
    text_color = (0, 0, 0)
    text_font = "Roboto"

    # ---------- log functions ---------- #
    log_file = None

    def start_log(self):
        if self.log_file is None: return
        name = self.log_file
        if os.path.exists(name):
            self.logbook = openpyxl.load_workbook(name)
            self.logsheet = self.logbook.worksheets[0]
            self.current_row = self.logsheet.max_row + 1
        else:
            self.logbook = openpyxl.workbook.Workbook()
            self.logsheet = self.logbook.worksheets[0]
            header = ["time", "rob x", "rob y", "rob z", "action id", "target x", "target x", "target x"]
            for col, item in enumerate(header, start=1):
                self.logsheet.cell(1, col, item)
            self.current_row = 2
        self.start_time = time.time()
        self.log("Start")

    def function_wrapper(self, func):
        def fun(x):
            self.log(func.__name__)
            func()

        return fun

    def log_clock(self, dt):
        self.log("Clock")

    def end_log(self):
        self.log("End")
        self.save_log()

    def save_log(self, *args):
        self.logbook.save(self.log_file)

    def log(self, id):
        if not logging: return
        stuff = (time.time() - self.start_time,  # time
                 *(rob.getl()[:3] if robot else ('-', '-', '-')),  # position
                 id,  # action id
                 *(self.waypoint if self.waypoint is not None else ('-', '-', '-'))  # waypoint target
                 )
        for col, item in enumerate(stuff, start=1):
            self.logsheet.cell(self.current_row, col, item)
        self.current_row += 1

    # ---------- util functions ---------- #
    @staticmethod
    def logDebug(*args):  # can remove return to log actions
        if debugging:
            print(args)

    @staticmethod
    def dimensions():
        return int(Config.get('graphics', 'width')) * (2 if mac else 1), \
               int(Config.get('graphics', 'height')) * (2 if mac else 1)

    @staticmethod
    def pos_hint_from_tuple(tup):
        return {'x': tup[0], 'y': tup[1]}

    def make_button(self, press, text="", pos=(0, 0), size=None, font_size=20, background_normal=None, border=None):
        if size is None: size = self.button_size_1
        pos = self.pos_hint_from_tuple(pos)
        if background_normal is None:  # if labeled box
            b = Button(text=text, pos_hint=pos, font_name=self.text_font, font_size=font_size, size=size,
                       size_hint=(None, None))
            b.background_color = self.button_color + (1,)
        else:  # if image
            b = Button(pos_hint=pos, size=size, size_hint=(None, None), background_normal=background_normal,
                       border=border)
        b.bind(on_press=self.function_wrapper(press))
        self.active_layout.add_widget(b)
        return b

    def make_label(self, text, pos, font):
        pos = self.pos_hint_from_tuple(pos)
        label = Label(text=text, pos_hint=pos, font_name=self.text_font, font_size=font, size_hint=(None, None),
                      halign="left", color=self.text_color + (1,))
        label.bind(texture_size=label.setter('size'))
        self.active_layout.add_widget(label)
        return label

    def make_slider(self, pos, function, range):
        pos = self.pos_hint_from_tuple(pos)
        min, max, default = range
        slider = Slider(pos_hint=pos, min=min, max=max, value=default, on_touch_move=function, on_touch_up=function,
                        size_hint=self.slider_size)
        self.active_layout.add_widget(slider)
        return slider

    # --------------------- init --------------------- #
    def build(self):
        self.start_log()
        self.root = self.active_layout = FloatLayout()

        self.build_sizes()

        # ----- session control ----- #
        self.build_session_control()

        # ----- camera ----- #
        self.build_cams()

        # ----- supervisory ----- #
        self.build_supervisory()

        # ----- nudge controls ----- #
        self.build_nudge()

        with self.root.canvas.before:
            Color(self.background_color[0], self.background_color[1], self.background_color[2], 1)
            self.rect = Rectangle(
                size=(int(Config.get('graphics', 'width')) * 2, int(Config.get('graphics', 'height')) * 2),
                pos=self.root.pos)

        return self.root

    def build_sizes(self):
        w, h = self.dimensions()
        w, h = w // 2, h // 2

        self.title_font = int(w / 20)
        self.slider_font = int(w / 40)

        i_w, i_h = (640, 480)
        self.main_img_size = (i_w, i_h)
        self.secondary_img_size = (int(i_w / 2), int(i_h / 2))

        self.button_size_1 = 1 / 8 * w, 1 / 8 * h
        self.slider_size = 0.18, 0.1  # this is a hint

        self.session_button_size = self.button_size_1
        self.camera_button_size = self.button_size_1[0] * 1.5, self.button_size_1[1]
        self.supervisory_button_size = self.button_size_1[0] * 2.25, self.button_size_1[1]
        self.elevation_button_size = self.button_size_1[0] * 1.2, self.button_size_1[1]
        self.gripper_button_size = self.elevation_button_size  # self.button_size_1[0] * 1.2, self.button_size_1[1] * 1.5

        self.rotation_button_size = 1 / 8 * w, 1 / 4 * h
        self.left_right_size = 1 / 8 * w, 1 / 10 * h
        self.forward_backward_size = 1 / 10 * w, 1 / 10 * h

        # slider range - min, max, default
        self.rotation_range = 1, 10, 5
        self.movement_range = 0.5, 5, 2

        # slider magnetic points
        self.rotation_magnet = 0.1
        self.move_magnet = 0.1

    def build_session_control(self):
        s_x, s_y = 0.05, 0.15
        self.session_label = self.make_label("Session Control", (s_x, s_y), font=self.title_font)
        self.session_start = self.make_button(self.on_session_start, text="Start", size=self.session_button_size,
                                              pos=(s_x, s_y - .1))
        self.session_stop = self.make_button(self.on_stop, text="Stop", size=self.session_button_size,
                                             pos=(s_x + .1, s_y - .1))

    def build_cams(self):
        self.establish_cameras()
        c_x, c_y = 0.2, 0.3
        self.switchCamButton = self.make_button(self.switch_camera, text="Switch View", pos=(c_x, c_y - 0.025),
                                                size=self.camera_button_size)
        pos_main = self.pos_hint_from_tuple((0.05, c_y + 0.08))
        pos_secondary = self.pos_hint_from_tuple((0.3, 0.05))
        self.main_img = Image(on_touch_down=self.on_press_image, pos_hint=pos_main, size=self.main_img_size,
                              source="resources/main.png", size_hint=(None, None))
        self.secondary_img = Image(pos_hint=pos_secondary, size=self.secondary_img_size,
                                   source="resources/secondary.png", size_hint=(None, None))
        self.root.add_widget(self.main_img)
        self.root.add_widget(self.secondary_img)
        Clock.schedule_interval(self.update_img, 1 / 60)

    def build_supervisory(self):
        su_x, su_y = 0.6, 0.6

        su_x1, su_x2 = su_x + .05, su_y + 0.2
        su_y1, su_y2 = su_y + 0.1, su_y + 0.2
        self.supervisoryLabel = self.make_label("Supervisory Control", (su_x, su_y + 0.3), font=self.title_font)
        self.returnButton = self.make_button(self.on_press_return, size=self.supervisory_button_size,
                                             text="Go to start position", pos=(su_x2, su_y1))
        self.approachButton = self.make_button(self.on_press_approach, text="Approach",
                                               size=self.supervisory_button_size, pos=(su_x1, su_y2))
        self.pauseButton = self.make_button(self.delete_thread, text="Delete", size=self.supervisory_button_size,
                                            pos=(su_x2, su_y2))
        self.pauseButton = self.make_button(self.on_press_pauseButton, text="Pause", size=self.supervisory_button_size,
                                            pos=(su_x1, su_y1))

    def build_nudge(self):
        n_x, n_y = self.supervisoryLabel.pos_hint['x'], 0
        self.nudgeLabel = self.make_label("Nudge Control", (n_x, n_y + 0.55), font=self.title_font)
        self.build_move()
        self.build_gripper()

    def build_move(self):
        border = (0, 0, 0, 0)
        m_x, m_y = tuple(self.nudgeLabel.pos_hint.values())

        self.moveSlider = self.make_slider((m_x, m_y - 0.15), self.on_move_slider, self.movement_range)
        self.moveSensitvityLabel = self.make_label(pos=(m_x, m_y - 0.075), font=self.slider_font,
                                                   text=("Arm Movement          %.1f cm" % self.moveSlider.value))
        self.nudgeLimit1 = self.make_label("0.5 cm", (m_x, m_y - 0.135), font=self.slider_font)
        self.nudgeLimit2 = self.make_label("5.0 cm", (m_x + .15, m_y - 0.135), font=self.slider_font)

        w1, h1 = self.left_right_size
        w2, h2 = self.forward_backward_size
        w, h = self.dimensions()
        d_long_x, d_short_x = w1 / w, w2 / w
        d_long_y, d_short_y = h2 / h, h1 / h

        left_right_y = m_y - .3
        forward_backward_x = m_x + .08 - d_short_x / 4

        # print(d_long*int(Config.get('graphics', 'width')))
        self.leftButton = self.make_button(self.on_press_leftButton,
                                           background_normal="resources/left.png", size=self.left_right_size,
                                           pos=(forward_backward_x - d_long_x, left_right_y), border=border)
        self.rightButton = self.make_button(self.on_press_rightButton,
                                            background_normal="resources/right.png", size=self.left_right_size,
                                            pos=(forward_backward_x + d_short_x, left_right_y), border=border)
        self.backwardButton = self.make_button(self.on_press_backwardButton,
                                               background_normal="resources/backward.png",
                                               size=self.forward_backward_size,
                                               pos=(forward_backward_x, left_right_y - d_long_y),
                                               border=border)
        self.forwardButton = self.make_button(self.on_press_forwardButton,
                                              background_normal="resources/forward.png",
                                              size=self.forward_backward_size,
                                              pos=(forward_backward_x, left_right_y + d_short_y),
                                              border=border)

        far_left = m_x
        far_right = (self.rightButton.pos_hint['x'] + d_long_x)
        middle = (far_left + far_right) / 2
        half = middle - far_left
        width = self.elevation_button_size[0] / self.dimensions()[0]
        pad = (half - width) / 2

        elevation_y = m_y - .5
        self.upButton = self.make_button(self.on_press_upButton, text="Lift",
                                         pos=(far_left + pad, elevation_y), size=self.elevation_button_size)
        self.downButton = self.make_button(self.on_press_downButton,
                                           text="Lower", pos=(middle + pad, elevation_y),
                                           size=self.elevation_button_size)

    def build_gripper(self):
        g_x, g_y = self.nudgeLabel.pos_hint['x'] + .2, self.nudgeLabel.pos_hint['y']
        open_close_y = self.upButton.pos_hint['y']

        self.gripperButtonOpen1 = self.make_button(self.on_press_gripperButtonOpen, text="Open \nGripper",
                                                   size=self.gripper_button_size, pos=(g_x, open_close_y))

        self.gripperButtonClose1 = self.make_button(self.on_press_gripperButtonClose, text="Close \nGripper",
                                                    size=self.gripper_button_size, pos=(g_x + .1, open_close_y))

        width = self.dimensions()[0] // 2
        far_right = self.gripperButtonClose1.size[0] / width / 2 + self.gripperButtonClose1.pos_hint['x']
        center_x = (g_x + far_right) / 2
        button_width = self.rotation_button_size[0] / width / 2
        offset = 0.005

        rotate_y = self.leftButton.pos_hint['y']
        border = (0, 0, 0, 0)
        self.rotateCounterClock = self.make_button(self.on_rotate_counter_clock,
                                                   background_normal="resources/counter_clock.png",
                                                   size=self.rotation_button_size,
                                                   pos=(center_x - button_width - offset, rotate_y), border=border)
        self.rotateClock = self.make_button(self.on_rotate_clock,
                                            background_normal="resources/clock.png", size=self.rotation_button_size,
                                            pos=(center_x + offset, rotate_y), border=border)

        self.rotationSlider = self.make_slider((g_x, g_y - 0.15), self.on_rotate_slider, self.rotation_range)
        self.rotationSensitivityLabel = self.make_label(pos=(g_x, g_y - 0.075), font=self.slider_font,
                                                        text="Rotation Angle          %.1f deg" % self.rotationSlider.value)
        self.nudgeLimit3 = self.make_label("1.0 deg", (g_x, g_y - 0.135), font=self.slider_font)
        self.nudgeLimit4 = self.make_label("10 deg", (g_x + .15, g_y - 0.135), font=self.slider_font)

    # --------------------- utils --------------------- #
    def get_nudge_move(self):
        return self.moveSlider.value / 100  # convert cm to meters

    def get_nudge_rotate(self):
        return math.radians(self.rotationSlider.value)

    def get_depth_scale(self):
        return self.depth_scale

    def on_session_start(self):
        self.logDebug("starting session")
        print(self.log_file)
        Clock.schedule_interval(self.log_clock, 1 / 10)
        Clock.schedule_interval(self.save_log, 2)  # save log every 2 seconds in case of error

    def on_stop(self):
        self.logDebug("Stopping")
        self.running = False
        if robot: rob.close()
        self.end_log()
        sys.exit()

    def _update_rect(self, instance, value):
        self.rect.pos = instance.pos
        self.rect.size = instance.size

    # --------------------- camera --------------------- #
    def establish_cameras(self, new=True):
        if camera:
            if new:
                self.pipeline1 = rs.pipeline()
                self.pipeline2 = rs.pipeline()
            profile1 = self.pipeline1.start(  # outer camera
                create_config(device_1, self.main_img_size if self.back_main_flag else self.secondary_img_size,
                              self.back_main_flag))
            profile2 = self.pipeline2.start(  # arm camera
                create_config(device_2, self.main_img_size if not self.back_main_flag else self.secondary_img_size,
                              not self.back_main_flag))  # always true for mask
            self.depth_scale = (
                profile1 if self.back_main_flag else profile2).get_device().first_depth_sensor().get_depth_scale()

    def switch_camera(self):
        self.logDebug("Swapping cameras")
        self.back_main_flag = not self.back_main_flag
        if camera:
            self.pipeline1.stop()
            self.pipeline2.stop()
        self.establish_cameras(new=False)
        self.main_img.reload()
        self.secondary_img.reload()
        self.delete_flag = True

    def on_press_image(self, instance, touch):
        self.delete_flag = False
        IMAGE_X, IMAGE_Y = self.main_img.pos
        view1_x, view1_y = self.main_img.size
        mouseX, mouseY = self.main_img.to_widget(*touch.pos)
        # self.logDebug("Mouse Click: ", mouseX, mouseY)
        if IMAGE_X < mouseX < IMAGE_X + view1_x and IMAGE_Y < mouseY < IMAGE_Y + view1_y:  # this doesnt work anymore
            if robot: Thread(target=self.waypoint_thread, args=[mouseX, mouseY]).start()
        else:
            self.logDebug("Mouse click out of range")
        self.log("on_press_image")

        self.main_img.reload()
        self.secondary_img.reload()

    # --------------------- overlays --------------------- #
    msg_time = None
    msg = None
    border_color = None  # clear

    def arm_median_depth(self, frame):
        if not camera: return
        depth_img = np.asanyarray(rs.video_stream_profile(frame.get_depth_frame().profile).get_intrinsics().get_data())
        w, h = depth_img.shape
        cropped = depth_img[w // 4: w // 4 * 3, h // 4: h // 4 * 3]
        return np.mean(cropped)

    @staticmethod
    def camera_mask(img):
        if not camera: return img
        alpha = .3

        table_height = -.2
        cam_to_grip = .135
        depth = rob.getl()[2] - table_height if robot else 1
        depth += cam_to_grip

        w, h, d = img.shape

        overlay = create_mask(depth, (w, h))

        overlay = cv2.resize(overlay, (h, w))
        mask = overlay[:, :, 3] / 255.0
        mask = cv2.merge((mask, mask, mask))
        overlay = cv2.cvtColor(overlay.astype('uint8'), cv2.COLOR_BGRA2BGR)

        img = img.astype('uint8')

        adjusted_overlay = overlay * alpha
        adjusted_img = img * (1 - mask * alpha)
        final_img = adjusted_img + adjusted_overlay
        # cv2.addWeighted(overlay, alpha, img, 1-alpha, 0, img)
        return final_img

    def add_notice(self, msg, t=3):
        self.msg = msg
        self.msg_time = time.time() + t

    def draw_notices(self, img):
        border_size = 8
        if self.border_color is not None:
            if self.border_color == (0, 0, 255):  # red
                cv2.rectangle(img, (0, 0), self.main_img_size, color=self.border_color, thickness=border_size * 2)
            else:
                img = border(img, color=self.border_color, thickness=border_size)
            # cv2.rectangle(img, (0, 0), self.main_img_size, color=self.border_color, thickness=border_size)
        if self.msg is not None:
            img = cv2.putText(img, str(self.msg), (10, 50), cv2.cv2.FONT_HERSHEY_PLAIN, 2.5, (0, 255, 255), 3)
            if time.time() > self.msg_time:
                self.msg = None
        return img

    # --------------------- supervisory --------------------- #
    def on_press_pauseButton(self):
        self.logDebug("Pausing supervisory")
        if robot: rob.stopl()
        self.delete_flag = False

    def delete_thread(self):
        self.delete_flag = True

    def on_press_approach(self):
        self.logDebug("approach")
        if robot: Thread(target=self.waypoint_thread_execute).start()

    def start_thread(self):
        self.logDebug("stop supervisory")
        rob.stopl()

    def on_press_stopButton(self):
        self.logDebug("stop supervisory")
        Thread(target=self.stop_thread).start()

    def stop_thread(self):  # add in a stop flag in the log file
        self.delete_flag = True
        Thread(target=self.home_thread).start()

    def on_press_return(self):
        self.logDebug("returning home")
        Thread(target=self.home_thread).start()

    def home_thread(self):
        if not robot: return
        rob_x, rob_y, rob_z, rob_rx, rob_ry, rob_rz = rob.getl()
        if rob_z < 0.21:
            rob.movels([(rob_x, rob_y, rob_z + 0.09, rob_rx, rob_ry, rob_rz),
                        (0.5, 0.4, 0.3, -2.849507334119959, 1.2795477952223713, -0.003164266628449624)], a, v)
        else:
            rob.movel((0.5, 0.4, 0.3, -2.849507334119959, 1.2795477952223713, -0.003164266628449624), a, v)

        time.sleep(0.2)
        self.delete_flag = True

    # --------------------- gripper --------------------- #
    def on_press_gripperButtonOpen(self):
        self.logDebug("open gripper")
        Thread(target=self.gripper_thread_open).start()

    def on_press_gripperButtonClose(self):
        self.logDebug("close gripper")
        Thread(target=self.gripper_thread_close).start()

    def gripper_thread_open(self):
        self.logDebug("Opening Gripper")
        if robot: robotiqgrip.open_gripper()

    def gripper_thread_close(self):
        self.logDebug("Close Gripper")
        if robot: robotiqgrip.close_gripper()

    def on_rotate_clock(self):
        Thread(target=self.rotate_thread, args=['rz-']).start()

    def on_rotate_counter_clock(self):
        Thread(target=self.rotate_thread, args=['rz+']).start()

    # --------------------- waypoints --------------------- #
    def waypoint_thread(self, mouseX, mouseY):  # no waypoints
        IMAGE_X, IMAGE_Y = self.main_img.pos
        view1_x, view1_y = self.main_img.size
        c, r = mouseX - IMAGE_X, view1_y - (mouseY - IMAGE_Y)
        depth_value = self.depth_image[int(r)][int(c)] * self.get_depth_scale()
        depth_point_initial = rs.rs2_deproject_pixel_to_point(self.aligned_depth_intrinsics, [c, r], depth_value)
        camera_coordinates_initial = np.array(depth_point_initial)
        print("Initial target=", camera_coordinates_initial)
        if depth_value == 0:  # todo: test this
            r, c = min(zip(*np.nonzero(self.depth_image)), key=lambda i: math.sqrt((r - i[0]) ** 2 + (c - i[1]) ** 2))
            depth_value = self.depth_image[int(r)][int(c)] * self.get_depth_scale()
            self.add_notice("Invalid target")
        depth_point = rs.rs2_deproject_pixel_to_point(self.aligned_depth_intrinsics, [c, r], depth_value)
        camera_coordinates = np.array(depth_point)
        # add if loop here to ensure usable camera coordinates
        print("Final target=", camera_coordinates)
        self.waypoint_pixel = int(c), int(r)  # todo: change this to allow work with both images
        if robot:
            if self.back_main_flag:
                camera_coordinates[2] = camera_coordinates[2] - 0.16
                camera_coordinates[0] = camera_coordinates[0] - 0.005
                self.waypoint = np.dot(R_back, camera_coordinates.T).T + t_back  # coding: utf-8
            if not self.back_main_flag:
                camera_coordinates[0] = camera_coordinates[0] - 0.02
                # coordinates in klampt's end effector frame
                EE_coordinates_K = np.dot(R_on_arm, camera_coordinates.T).T + t_on_arm

                # coordinate in the ur5 default end effector frame
                EE_coordinates_ur5 = np.dot(R_k2ur5, EE_coordinates_K.T).T

                # tranformation matrix from the end effector to robot base
                Transform_EE2BASE = rob.get_pose()
                T_EE2BASE = Transform_EE2BASE.array

                # rotation matrix
                R_EE2BASE = np.array([[T_EE2BASE[0, 0], T_EE2BASE[0, 1], T_EE2BASE[0, 2]],
                                      [T_EE2BASE[1, 0], T_EE2BASE[1, 1], T_EE2BASE[1, 2]],
                                      [T_EE2BASE[2, 0], T_EE2BASE[2, 1], T_EE2BASE[2, 2]]])
                # translation
                Tr_EE2BASE = np.array([T_EE2BASE[0, 3], T_EE2BASE[1, 3], T_EE2BASE[2, 3]])

                # from ur5 end effector frame to ur5 base frame
                robot_coordinates = np.dot(R_EE2BASE, EE_coordinates_ur5.T).T + Tr_EE2BASE
                # r_x, r_y, r_z = robot_coordinates[0], robot_coordinates[1], robot_coordinates[2]
                # adjustment for gripper length
                # r_z = r_z+0.11
                robot_coordinates[2] = robot_coordinates[2] + 0.14

                # self.waypoint = np.array([r_x, r_y, r_z])
                self.waypoint = robot_coordinates
        else:
            self.logDebug("Depth Value", depth_value, camera_coordinates)

    def get_image_coordinates(self, mouseX, mouseY):
        IMAGE_X, IMAGE_Y = self.main_img.pos
        view1_x, view1_y = self.main_img.size
        return mouseX - IMAGE_X, view1_y - (mouseY - IMAGE_Y)

    def get_depth(self, pos):
        c, r = self.get_image_coordinates(pos[0], pos[1])
        return self.depth_image[int(r)][int(c)] * self.get_depth_scale()

    def on_press_waypointButtonExecute(self):
        pass
        # Thread(target=self.waypoint_thread_execute).start()

    def waypoint_thread_execute(self):
        if not self.delete_flag:
            if robot and self.waypoint is not None:
                x, y, z = self.waypoint
                if x == -0.23073204125 and y == 0.26365954534999997 and z == 0.022712601449999998:
                    bad_target = True
                else:
                    bad_target = False

                if self.back_main_flag:
                    # check if x,y,z within safe workspace
                    if (x ** 2 + y ** 2 + z ** 2) < .8 ** 2 and z > -.2 and (x ** 2 + y ** 2) > .01 and not bad_target:
                        rob.movels(
                            [(x + 0.025, y + 0.025, z + 0.09, 3, -1, 0), (x + 0.025, y + 0.025, z - 0.02, 3, -1, 0)], a,
                            v)
                        # rob.movel((x + 0.025, y + 0.025, z - 0.02, 3, -1, 0), a, v)
                        self.border_color = None
                    else:
                        self.waypoint_error()

                if not self.back_main_flag:
                    # check if x,y,z within safe workspace
                    if (x ** 2 + y ** 2 + z ** 2) < .8 ** 2 and z > -.2 and (x ** 2 + y ** 2) > .01 and not bad_target:
                        rob.movels([(x, y, z + 0.09, 3, -1, 0), (x, y, z, 3, -1, 0)], a, v)
                        # rob.movel((x, y, z, 3, -1, 0), a, v)
                        self.border_color = None
                    else:
                        self.waypoint_error()

            time.sleep(0.2)
            self.delete_flag = True

    def waypoint_error(self):
        self.border_color = (0, 0, 255)
        self.add_notice("Target outside workspace")

    def nudge_error(self):
        self.border_color = (0, 0, 255)
        self.add_notice("Target outside workspace")

    def nudge_close(self):
        self.border_color = (0, 255, 255)
        self.add_notice("Close to workspace edge")

    def draw_waypoint(self, img):
        if self.waypoint_pixel is None: return img
        radius = 5
        color = (0, 0, 255)  # red in BGR
        return cv2.circle(img, self.waypoint_pixel, radius, color, -1)

    # --------------------- sliders --------------------- #
    def on_move_slider(self, slider, *args):
        val = slider.value
        self.moveSensitvityLabel.text = "Arm Movement          %.1f cm" % good_round(val, 0.5)

    def on_rotate_slider(self, slider, *args):
        val = slider.value
        self.rotationSensitivityLabel.text = "Rotation Angle          %.1f deg" % good_round(val, 0.5)

    # --------------------- movement --------------------- #
    def on_press_upButton(self):
        Thread(target=self.move_thread, args=["up"]).start()

    def on_press_downButton(self):
        Thread(target=self.move_thread, args=["down"]).start()

    def on_press_leftButton(self):
        Thread(target=self.move_thread, args=["left"]).start()

    def on_press_rightButton(self):
        Thread(target=self.move_thread, args=["right"]).start()

    def on_press_backwardButton(self):
        Thread(target=self.move_thread, args=["backward"]).start()

    def on_press_forwardButton(self):
        Thread(target=self.move_thread, args=["forward"]).start()

    # --------------------- general action functions --------------------- #
    def move_thread(self, direction):
        self.border_color = None
        if not robot:
            return

        Transform_EE2BASE_move = rob.get_pose()

        T_EE2BASE_move = Transform_EE2BASE_move.array
        R_EE2BASE_move = np.array(
            [[T_EE2BASE_move[0, 0], T_EE2BASE_move[0, 1], T_EE2BASE_move[0, 2]],  # T_EE2BASE_move[:3, :3]
             [T_EE2BASE_move[1, 0], T_EE2BASE_move[1, 1], T_EE2BASE_move[1, 2]],
             [T_EE2BASE_move[2, 0], T_EE2BASE_move[2, 1], T_EE2BASE_move[2, 2]]])
        # translation
        Tr_EE2BASE_move = np.array(
            [T_EE2BASE_move[0, 3], T_EE2BASE_move[1, 3], T_EE2BASE_move[2, 3]])  # T_EE2BASE_move[:, 3]

        T_BASE2EE_move = inv(T_EE2BASE_move)
        # rotation
        R_BASE2EE_move = np.array([[T_BASE2EE_move[0, 0], T_BASE2EE_move[0, 1], T_BASE2EE_move[0, 2]],
                                   [T_BASE2EE_move[1, 0], T_BASE2EE_move[1, 1], T_BASE2EE_move[1, 2]],
                                   [T_BASE2EE_move[2, 0], T_BASE2EE_move[2, 1], T_BASE2EE_move[2, 2]]])
        # translation
        Tr_BASE2EE_move = np.array([T_BASE2EE_move[0, 3], T_BASE2EE_move[1, 3], T_BASE2EE_move[2, 3]])

        rob_x, rob_y, rob_z, rob_rx, rob_ry, rob_rz = rob.getl()
        rob_pos_in_base = np.array([rob_x, rob_y, rob_z])
        rob_pos_in_EE_urx = np.dot(R_BASE2EE_move, rob_pos_in_base.T).T + Tr_BASE2EE_move
        rob_pos_EE_k = np.dot(R_k2ur5_inv, rob_pos_in_EE_urx.T).T
        rob_pos_in_cam = np.dot(R_on_arm_inv, rob_pos_EE_k.T).T + t_on_arm_inv

        cam_x, cam_y, cam_z = rob_pos_in_cam

        if direction == "up":
            camera_pos = np.array([cam_x, cam_y, cam_z - self.get_nudge_move()])

        if direction == "down":
            camera_pos = np.array([cam_x, cam_y, cam_z + self.get_nudge_move()])

        if direction == "left":
            camera_pos = np.array([cam_x - self.get_nudge_move(), cam_y, cam_z])

        if direction == "right":
            camera_pos = np.array([cam_x + self.get_nudge_move(), cam_y, cam_z])

        if direction == "forward":
            camera_pos = np.array([cam_x, cam_y - self.get_nudge_move(), cam_z])

        if direction == "backward":
            camera_pos = np.array([cam_x, cam_y + self.get_nudge_move(), cam_z])

        # rob_pos = np.dot(R_back, camera_pos.T).T + t_back
        rob_target_EE_K = np.dot(R_on_arm, camera_pos.T).T + t_on_arm
        rob_target_EE_urx = np.dot(R_k2ur5, rob_target_EE_K.T).T
        rob_target_base = np.dot(R_EE2BASE_move, rob_target_EE_urx.T).T + Tr_EE2BASE_move
        r_x, r_y, r_z = rob_target_base[0], rob_target_base[1], rob_target_base[2]
        self.logDebug(rob_target_base)
        if (r_x ** 2 + r_y ** 2 + r_z ** 2) < .8 ** 2 and r_z > -.2 and (r_x ** 2 + r_y ** 2) > .01:
            if (r_x ** 2 + r_y ** 2 + r_z ** 2) > .78 ** 2 or r_z < -.18 or (r_x ** 2 + r_y ** 2) < .18 ** 2:
                self.nudge_close()
            else:
                self.border_color = None
            rob.movel((r_x, r_y, r_z, rob_rx, rob_ry, rob_rz), a, v)

        else:
            self.nudge_error()

    def rotate_thread(self, direction):
        if not robot: return
        rob_o = rob.get_orientation()
        rob_pos = rob.getl()
        rob_rx, rob_ry, rob_rz = rob_pos[3], rob_pos[4], rob_pos[5]

        acc = 0.1
        vel = 0.5

        if direction == "rz+":
            rob_o.rotate_zb(rob_rz + self.get_nudge_rotate())
            rob.set_orientation(rob_o, acc, vel)

        elif direction == "rz-":
            rob_o.rotate_zb(rob_rz - self.get_nudge_rotate())
            rob.set_orientation(rob_o, acc, vel)

        # else:
        #     print("TBD")

    def update_img(self, *args):
        # kivy's recommendation for video streaming:
        # https://kivyapps.wordpress.com/video-streaming-using-kivy-and-python/
        if not camera:
            ret, outer = video.read()
            arm = np.ones(self.secondary_img_size + (3,)) * 150
        else:
            # Get frameset of color and depth
            frames1 = self.pipeline1.wait_for_frames()
            frames2 = self.pipeline2.wait_for_frames()
            # frames.get_depth_frame() is a 640x360 depth image

            # Align the depth frame to color frame
            aligned_frames1 = align.process(frames1)
            aligned_frames2 = align.process(frames2)

            # Get aligned frames
            color_frame1 = aligned_frames1.get_color_frame()
            color_frame2 = aligned_frames2.get_color_frame()

            color_image1 = np.asanyarray(color_frame1.get_data())
            color_image2 = np.asanyarray(color_frame2.get_data())

            if self.back_main_flag:
                aligned_depth_frame = aligned_frames1.get_depth_frame()
            else:
                aligned_depth_frame = aligned_frames2.get_depth_frame()

            self.aligned_depth_intrinsics = rs.video_stream_profile(aligned_depth_frame.profile).get_intrinsics()
            self.depth_image = np.asanyarray(aligned_depth_frame.get_data())

            outer = color_image1
            arm = color_image2

        if camera: arm = self.camera_mask(arm)
        # outer = self.draw_waypoint(outer)

        if not self.delete_flag:
            if self.back_main_flag:
                frame_big = outer
                frame_small = arm
                outer = self.draw_waypoint(outer)
            if not self.back_main_flag:
                frame_big = arm
                frame_small = outer
                arm = self.draw_waypoint(arm)
        else:
            if self.back_main_flag:
                frame_big = outer
                frame_small = arm
            if not self.back_main_flag:
                frame_big = arm
                frame_small = outer
        # frame_small = outer if not self.back_main_flag else arm  # cv2.resize(outer if not self.back_main_flag else arm, self.secondary_img_size)

        frame_big = self.draw_notices(frame_big)
        cv2.imwrite("resources/main.png", frame_big)
        cv2.imwrite("resources/secondary.png", frame_small)
        self.main_img.reload()
        self.secondary_img.reload()
        # cv2.imshow('Fig1',frame_big)
        # cv2.imshow('Fig2',frame_small)


def create_mask(depth, shape):
    h, w = shape
    img = np.zeros((h, w, 4))
    red = (0, 0, 255, 255)  # bgr

    # camera shadow
    fov_h = 64  # 86 - https://www.intel.com/content/www/us/en/support/articles/000030385/emerging-technologies/intel-realsense-technology.html
    fov_v = 41  # 57

    cam_width = .098
    cam_height = .024

    # camera stuff
    shadow_theta_h = math.degrees(math.atan(cam_width / 2 / depth))
    shadow_theta_v = math.degrees(math.atan(cam_height / 2 / depth))

    shadow_w = w * (shadow_theta_h / fov_h)
    shadow_h = h * (shadow_theta_v / fov_v)
    cx, cy = w // 2, h // 2
    cv2.rectangle(img, (int(cx - shadow_w), int(cy - shadow_h)), (int(cx + shadow_w), int(cy + shadow_h)), color=red,
                  thickness=-1)

    # gripper stuff
    gripper_width = 0.11
    gripper_length = 0.022
    offset = 0.09
    # cam_to_grip = .135

    cam_dh = int(w * math.degrees(math.atan(gripper_width / 2 / depth)) / fov_h)
    cam_dv = int(h * math.degrees(math.atan(gripper_length / depth)) / fov_v)
    offset_v = 0.09
    offset_h = 0.01
    # cam_to_grip = .135

    cam_dh = int(w * math.degrees(math.atan(gripper_width / 2 / depth)) / fov_h)
    cam_dv = int(h * math.degrees(math.atan(gripper_length / depth)) / fov_v)
    offset_v = int(h * math.degrees(math.atan(offset_v / depth)) / fov_v)
    offset_h = int(w * math.degrees(math.atan(offset_h / depth) / fov_h))
    width_h = 10

    cv2.rectangle(img, (cx - cam_dh + offset_h, cy + offset_v),
                  (cx - cam_dh - width_h + offset_h, cy + offset_v + cam_dv), red, -1)
    cv2.rectangle(img, (cx + cam_dh + offset_h, cy + offset_v),
                  (cx + cam_dh + width_h + offset_h, cy + offset_v + cam_dv), red, -1)
    # cv2.imwrite("test.png", img)
    return img


def create_config(device, size, depth):
    config = rs.config()
    config.enable_device(device)
    print(size)
    if depth:
        config.enable_stream(rs.stream.depth, size[0], size[1], rs.format.z16, 30)
    config.enable_stream(rs.stream.color, size[0], size[1], rs.format.bgr8, 30)
    return config


def border(img, color=(255, 0, 0), thickness=10, length=30, spacing=10):
    h, w = img.shape[:2]
    x = y = 0
    while x < w:
        far = min(w, x + length)
        cv2.rectangle(img, (x, 0), (far, thickness), color, thickness=-1)
        cv2.rectangle(img, (x, h), (far, h - thickness), color, thickness=-1)
        x += length + spacing
    while y < h:
        far = min(h, y + length)
        cv2.rectangle(img, (0, y), (thickness, far), color, thickness=-1)
        cv2.rectangle(img, (w, y), (w - thickness, far), color, thickness=-1)
        y += length + spacing
    # print(img.shape)
    return img


import os

os.environ['KIVY_GL_BACKEND'] = 'sdl2'  # idk what this is
from kivy.config import Config

# --------------------- configs --------------------- #
# 0 being off 1 being on as in true / false
# you can use 0 or 1 && True or False
Config.set('graphics', 'resizable', '0')

# fix the width of the window
# was 1500 * 900  # actually was 1600, 1000
Config.set('graphics', 'width', '640' if mac else '1280')
Config.set('graphics', 'height', '420' if mac else '840')

# acceleration and velocity
v = 0.5
a = 0.04

# start the robot arm
if not robot:
    rob = None
    robotiqgrip = None
else:
    # """
    rob = urx.Robot("192.168.1.105")
    rob.set_tcp((0, 0, 0.18, 0, 0, 0))
    time.sleep(0.2)
    robotiqgrip = Robotiq_Two_Finger_Gripper(rob)
    # """

# camera ids
device_1 = '911222060024'
device_2 = '916512060628'

# print("Depth Scale is: ", depth_scale2)

# Create an align object
# rs.align allows us to perform alignment of depth frames to others frames
# The "align_to" is the stream type to which we plan to align depth frames.
if camera:
    align_to = rs.stream.color
    align = rs.align(align_to)

# R = np.array([[0.4729613, -0.02498102, 0.88072899],
#               [-0.88017677, 0.03193417, 0.47357054],
#               [-0.03995563, -0.99917774, -0.00688409]])
# t = np.array([-0.43179749, 0.24159906, 0.15602798])

# inversed_R = np.array([[0.4729613, -0.88017677, -0.03995563],
#                        [-0.02498102, 0.03193417, -0.99917774],
#                        [0.88072899, 0.47357054, -0.00688409]])
# inversed_t = np.array([0.42310757, 0.13739768, 0.26695648])

# R_back = np.array([[0.4729613, -0.02498102, 0.88072899],
#                    [-0.88017677, 0.03193417, 0.47357054],
#                    [-0.03995563, -0.99917774, -0.00688409]])
# t_back = np.array([-0.43179749, 0.29159906, 0.15602798])

# inversed_R_back = np.array([[0.472961306628662, -0.880176777913776, -0.0399556246279741],
#                             [-0.0249810255046716, 0.0319341736638282, -0.999177741061022],
#                             [0.880728990275949, 0.473570537677987, -0.00688408697281732]])
# inversed_t_back = np.array([0.467116421543204, 0.135800965465922, 0.243277953925307])

R_back = np.array([[0.69632745, -0.0119904, 0.71762408],
                   [-0.71683597, -0.06134814, 0.69453769],
                   [0.03569711, -0.99804441, -0.05131354]])
t_back = np.array([-0.1913692, 0.29480225, 0.02032541])

inversed_R_back = np.array([[0.69632745, -0.71683597, 0.03569711],
                            [-0.0119904, -0.06134814, -0.99804441],
                            [0.71762408, 0.69453769, -0.05131354]])
inversed_t_back = np.array([0.34385492, 0.03607663, -0.06637716])

# transformation from camera space to end effector space, this end effector is from klampt
R_on_arm = np.array([[0.0330604524021490, -0.170386879727978, 0.984822480299637],
                     [-0.998170445145012, 0.0442810919530209, 0.0411697380907021],
                     [-0.0506237980200431, -0.984381783716060, -0.168611194652118]])
t_on_arm = np.array([-0.00032808, -0.01807384, 0.09210034])

R_on_arm_inv = np.array([[0.03306045240214919338, -0.99817044514501170206, -0.050623798020043130422],
                         [-0.170386879727978593, 0.04428109195302085153, -0.98438178371605963418],
                         [0.98482248029963678239, 0.041169738090701849354, -0.16861119465211731576]])

t_on_arm_inv = np.array([-0.013367457435318322066, 0.09140632581353858694, 0.01659634417369614205])

# transformation matrix from klampt definition to ur5 default
R_k2ur5 = np.array([[0, -1, 0], [0, 0, -1], [1, 0, 0]])
R_k2ur5_inv = inv(R_k2ur5)

# R = np.array([[0.57836741, -0.04958971,  0.81426777],
#         [-0.80568323,  0.12180494,  0.57968793],
#         [-0.1279284,  -0.99131449,  0.03049441]])
# t = np.array([-0.42742927,  0.18999053,  0.11613165])

# inversed_R = np.array([[0.578367411315431, -0.805683225801551, -0.127928392510784],
#                        [-0.049589714368207,0.121804945255018,-0.991314488298266],
#                        [0.814267766897599,0.579687923174278,0.030494409779250]])
# inversed_t = np.array([0.415139878796616,0.070785105677450,0.234365295307650])

if __name__ == '__main__':
    app = KinectView()
    app.run()
    app.end_log()
    cv2.destroyAllWindows()